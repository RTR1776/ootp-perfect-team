"""Turns scripts/.inventory.json into "PT Tourney Data Inventory.xlsx" at the repo root.
Run via `pnpm inventory`, which does the node step first. Recalc is not needed on the
Summary sheet's formulas for the numbers to be right in Excel, but run scripts/recalc.py
if you want the cached values populated for anything that reads the file programmatically."""
import json, os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
d = json.load(open(os.path.join(HERE, '.inventory.json'), encoding='utf-8'))
OUT = os.path.join(HERE, '..', '..', 'PT Tourney Data Inventory.xlsx')
TIERS = ['Diamond','Gold','Silver','Bronze','Iron','Open','Live','Cap','Other']
T = sorted(d['tourneys'], key=lambda r:(TIERS.index(r['tier']), 1 if r['retired'] else 0,
                                        0 if r['filesOnDisk']>0 else 1, -r['entries'], r['name']))
DR = sorted(d['drafts'], key=lambda r:(1 if r['retired'] else 0, 0 if r['filesOnDisk']>0 else 1,
                                       -r['entries'], r['name']))
TODAY = d['today']
F='Arial'; BOX=Border(bottom=Side(style='thin',color='C8D4D6'))
HDR=PatternFill('solid',fgColor='1F3F49')
TIERFILL={'Diamond':'E4ECF7','Gold':'FBF1DA','Silver':'EDEFF0','Bronze':'F6E7DC','Iron':'E9E9E7',
          'Open':'E7F1EC','Live':'F3E8F2','Cap':'EAF0F1','Other':'F2F2F0'}
HAVE=PatternFill('solid',fgColor='D8ECE0'); NONE=PatternFill('solid',fgColor='FBE3DC'); STALE=PatternFill('solid',fgColor='FCF0D8')
COLS=[('Tier','tier',9),('Tournament','name',42),('PT id','ptid',7),('Series (file name)','series',20),
      ('Runs on file','runsOnFile',9),('Files','filesOnDisk',7),('Latest run','latestRun',9),
      ('Latest date','latestDate',11),('Days old','daysBehind',8),('Missing last 7d','missingCount',10),
      ('Event ids still grabbable','grabIds',44),('Card cap','cap',9),('Field','field',7),
      ('My entries','entries',9),('My pts','pts',8),('Still running?','retired',34),('Run numbering','numbering',20)]

def build(ws, title, sub, rows):
    ws['A1']=title; ws['A1'].font=Font(name=F,size=14,bold=True,color='1F3F49')
    ws['A2']=sub;   ws['A2'].font=Font(name=F,size=9,italic=True,color='5A6E71')
    for j,(h,_,_) in enumerate(COLS,1):
        c=ws.cell(row=4,column=j,value=h); c.fill=HDR
        c.font=Font(name=F,size=9,bold=True,color='FFFFFF')
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    for i,r in enumerate(rows):
        for j,(_,k,_) in enumerate(COLS,1):
            v=r.get(k,'')
            if k=='retired': v = v if v else 'yes'
            c=ws.cell(row=5+i,column=j,value=v)
            c.font=Font(name=F,size=9.5); c.border=BOX
            if isinstance(v,(int,float)): c.alignment=Alignment(horizontal='right')
        t=ws.cell(row=5+i,column=1)
        t.fill=PatternFill('solid',fgColor=TIERFILL.get(r['tier'],'F2F2F0')); t.font=Font(name=F,size=9.5,bold=True)
        st=ws.cell(row=5+i,column=6)
        if r['filesOnDisk']==0: st.fill=NONE; st.font=Font(name=F,size=9.5,bold=True,color='9C3A1E')
        elif isinstance(r['daysBehind'],int) and r['daysBehind']>7: st.fill=STALE
        else: st.fill=HAVE
        if r['retired']: ws.cell(row=5+i,column=16).font=Font(name=F,size=9,italic=True,color='7A6A55')
        if str(r['numbering']).startswith('MISMATCH'):
            ws.cell(row=5+i,column=17).font=Font(name=F,size=9,bold=True,color='9C3A1E')
    for j,(_,_,w) in enumerate(COLS,1): ws.column_dimensions[get_column_letter(j)].width=w
    ws.freeze_panes='C5'; ws.row_dimensions[4].height=30
    ws.auto_filter.ref=f"A4:{get_column_letter(len(COLS))}{4+len(rows)}"
    return 4+len(rows)

wb=openpyxl.Workbook(); ws=wb.active; ws.title='Tournaments'
last_t=build(ws,'Tournaments — stat-export coverage',
  f'Diamond → Gold → Silver → Bronze → Iron → Open → Live → Cap → Other (card level wins over Cap/Live tags); inside each tier, events with data first, then by how often you play them, retired ones last. "Event ids still grabbable" = runs from the last 7 days you do NOT have, the only ones OOTP will still export. Generated {TODAY}.', T)
ws2=wb.create_sheet('Perfect Drafts')
last_d=build(ws2,'Perfect Drafts — stat-export coverage',
  f'Same columns. PD Daily is the biggest points category and the thinnest data. Generated {TODAY}.', DR)
ws3=wb.create_sheet('Summary')
ws3['A1']='Coverage summary'; ws3['A1'].font=Font(name=F,size=14,bold=True,color='1F3F49')
ws3['A2']=f'As of {TODAY}. OOTP only exports about the last 7 days, so any run older than that is gone for good.'
ws3['A2'].font=Font(name=F,size=9,italic=True,color='5A6E71')
srows=[('Tournaments listed',f'=COUNTA(Tournaments!B5:B{last_t})'),
 ('   …still running',f'=COUNTIF(Tournaments!P5:P{last_t},"yes")'),
 ('   …with at least one export',f'=COUNTIF(Tournaments!F5:F{last_t},">0")'),
 ('   …with an export inside 7 days',f'=COUNTIFS(Tournaments!F5:F{last_t},">0",Tournaments!I5:I{last_t},"<=7")'),
 ('Perfect Drafts listed',f"=COUNTA('Perfect Drafts'!B5:B{last_d})"),
 ('   …still running',f"=COUNTIF('Perfect Drafts'!P5:P{last_d},\"yes\")"),
 ('   …with at least one export',f"=COUNTIF('Perfect Drafts'!F5:F{last_d},\">0\")"),
 ('',''),
 ('Export files on disk',f"=SUM(Tournaments!F5:F{last_t})+SUM('Perfect Drafts'!F5:F{last_d})"),
 ('Runs imported to the database',f"=SUM(Tournaments!E5:E{last_t})+SUM('Perfect Drafts'!E5:E{last_d})"),
 ('Runs still grabbable (last 7 days)',f"=SUM(Tournaments!J5:J{last_t})+SUM('Perfect Drafts'!J5:J{last_d})")]
for i,(a,f) in enumerate(srows):
    ws3.cell(row=4+i,column=1,value=a).font=Font(name=F,size=10,bold=not a.startswith('   '))
    if f:
        c=ws3.cell(row=4+i,column=2,value=f); c.font=Font(name=F,size=10); c.alignment=Alignment(horizontal='right')
notes=['','Column notes',
 '• PT id — the first three digits of an event id. A run is <PT id><4-digit run>, so tournament 185 run 151 is event 1850151.',
 '• Still running? — OOTP REUSES an id slot when a tournament is renamed. Slot 218 was Treasure Trove until Aug 6 and is 6L Power Play now. A row naming a successor is retired: its runs cannot be exported and those ids belong to the new event.',
 '• Runs on file vs Files — "Runs on file" is what the database imported, "Files" is what sits in Archive/Completed. A gap means pnpm import:observed has not been run.',
 '• Days old — age of your newest export. Over 7 and the older runs are unrecoverable.',
 '• Event ids still grabbable — from each event\'s own id ladder and cadence in your results history. Blank means you have everything, the event is retired, or there is no ladder to compute from.',
 '• Run numbering — the filer names files by the id you type; where that counter disagrees with the PT run number the row says MISMATCH and the missing-run maths is unreliable.']
for i,t in enumerate(notes):
    c=ws3.cell(row=4+len(srows)+i,column=1,value=t)
    c.font=Font(name=F,size=10,bold=(t=='Column notes'),color='1F3F49' if t=='Column notes' else '333333')
ws3.column_dimensions['A'].width=46; ws3.column_dimensions['B'].width=14
wb.save(OUT); print('wrote', os.path.normpath(OUT), '|', len(T), 'tournaments,', len(DR), 'drafts')
