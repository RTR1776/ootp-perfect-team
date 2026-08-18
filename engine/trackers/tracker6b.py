from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
F="Arial"
hf=PatternFill("solid",fgColor="1F3B57"); hfont=Font(name=F,bold=True,color="FFFFFF",size=11)
title=Font(name=F,bold=True,size=15); sub=Font(name=F,italic=True,color="666666",size=10)
bf=Font(name=F,size=11); gray=Font(name=F,size=10,color="888888")
thin=Side(style="thin",color="CCCCCC"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal="center"); left=Alignment(horizontal="left")

# ALL 10 categories. Target = PTCS5 line x 5/4 (PTCS6 is 5 weeks).
CATS=[
 ("Iron","C",135,"—"),("Bronze","D",130,"—"),("Silver","E",129,"87 (missed 16)"),
 ("Gold","F",114,"119 ✅"),("Diamond","G",101,"283 ✅"),("Open","H",129,"14 (not chased)"),
 ("Live","I",176,"165 ✅"),("Cap","J",179,"171 ✅"),
 ("PD Daily","K",334,"403 ✅"),("PD Weekly","L",96,"17 (not chased)"),
]
DAYS=[
 ["Mon Aug 3","Mon",0,0,4,1,45,1,1,31,11,0,"HUGE: D&Friends Slots WINNER +25 Dia/+25 Cap; Diamond Heart 2nd +20 Dia; Meeting 3/4 +10 PDD; Deadball 9-16 +3S/+3C"],
 ["Tue Aug 4","Tue",0,0,6,3,10,0,0,6,3,0,"Deadball Slots 5-8 +6S/+6C; Diamond Heart 5-8 +10D; Early Gold 17-32 +3G; Struggling 9-16 +3P"],
 ["Wed Aug 5","Wed",0,0,4,0,1,3,0,2,0,0,"quiet - all 17-32s. Open OPENED +3 (Night of Living Deadball)"],
 ["Thu Aug 6","Thu",0,0,7,7,1,0,11,1,12,0,"BIG: Live Speed Run 3/4 +10L/+10P; Late Silver 9-16 +6S; Low Gold Retro 5-8 +6G; Sporer's Golden 17-32 +1G; Deadball Slots 17-32 +1S/+1C; Trying to Look Busy 17-32 +1P; Diamond Heart 33-64 +1D. Live All Night 17-32 +1L/+1P (earlier)"],
 ["Fri Aug 7","Fri",0,0,0,16,6,0,0,6,30,0,"BEST DAY: Nocturnal Mixed Bag WINNER +25 PDD; Low Gold Retro 3/4 +10G; Golden Heart 5-8 +6G; D&Friends Slots 9-16 +3D/+3C; Diamond Slots 9-16 +3D/+3C; Historical Primetime 9-16 +3P; Pizza for Dinner + Look Busy 17-32 +1P each. SILVER BLANKED. Live All Night 33-64 = 0. Thursday Overnight (PD Weekly) ELIMINATED - placement unknown, unscored"],
 ["Sat Aug 8","Sat",0,0,4,1,3,0,0,0,7,0,"Just Order Pizza 5-8 +6P; Late Silver 17-32 +3S; Diamond Heart 17-32 +3D; Slamboree 33-64 +1S; Golden Heart 17-32 +1G; Look Busy 17-32 +1P. PTCS 5 PLAYED (see PTCS-PTMS sheet)"],
 ["Sun Aug 9","Sun",0,0,13,1,4,0,0,3,10,1,"SILVER BREAKS OUT +13: Late Silver 5-8 (128) +10S; Silver Heart 9-16 +3S. Good Orderly Evening 3/4 +10P; Diamond Slots 9-16 +3D/+3C. FIRST PD WEEKLY POINT: Up and Down the Ladder 33-64 +1. Early Gold 33-64 +1G; Diamond Heart 17-32 +1D"],
 ["Mon Aug 10","Mon",0,0,0,0,0,0,0,0,0,0,"Struggling to Sleep 33-64 (64) = 0"],]
wb=Workbook()
log=wb.active; log.title="Daily Log"
log["A1"]="PTCS 6 — Daily Log (Aug 3 – Sep 6)"; log["A1"].font=title
log["A2"]="All 10 categories. 5-week period. Enter points earned per day."; log["A2"].font=sub
cols=["Date","Day","Iron","Bronze","Silver","Gold","Diamond","Open","Live","Cap","PD Daily","PD Weekly","Notes"]
for c,h in enumerate(cols,1):
    x=log.cell(4,c,h); x.font=hfont; x.fill=hf; x.alignment=ctr; x.border=bd
r=5
for row in DAYS:
    for c,v in enumerate(row,1):
        x=log.cell(r,c,v); x.font=bf; x.border=bd; x.alignment=left if c in(1,2,13) else ctr
    r+=1
for _ in range(40):
    for c in range(1,14): log.cell(r,c,"").border=bd
    r+=1
LAST=r-1
for c,w in enumerate([12,6,8,8,8,8,9,8,8,8,10,10,52],1): log.column_dimensions[get_column_letter(c)].width=w
log.freeze_panes="C5"

d=wb.create_sheet("Dashboard",0)
d["A1"]="PTCS 6 — All 10 Categories (Aug 3 – Sep 6)"; d["A1"].font=title
d["A2"]="5-week period. Targets = PTCS5 cutoff × 5/4 (estimates — replace with official once week-1 standings post)."; d["A2"].font=sub
head=["Category","Total","Target (est)","Gap","Per week","Per day","PTCS5 result","Status"]
for c,h in enumerate(head,1):
    x=d.cell(4,c,h); x.font=hfont; x.fill=hf; x.alignment=ctr; x.border=bd
rr=5
for name,col,tgt,p5 in CATS:
    d.cell(rr,1,name).font=Font(name=F,bold=True,size=11)
    d.cell(rr,2,f"=SUM('Daily Log'!{col}5:{col}{LAST})").font=Font(name=F,bold=True,size=11)
    d.cell(rr,3,tgt).font=bf
    d.cell(rr,4,f"=MAX(0,C{rr}-B{rr})").font=bf
    d.cell(rr,5,f"=ROUND(D{rr}/5,1)").font=bf
    d.cell(rr,6,f"=ROUND(D{rr}/35,1)").font=bf
    d.cell(rr,7,p5).font=gray
    d.cell(rr,8,f'=IF(B{rr}>=C{rr},"QUALIFIED","+"&TEXT(D{rr},"0")&" to go")')
    for c in range(1,9):
        x=d.cell(rr,c); x.border=bd
        if c not in (1,7,8): x.alignment=ctr
    rr+=1
for c,w in enumerate([13,9,13,8,10,9,17,16],1): d.column_dimensions[get_column_letter(c)].width=w

nr=rr+2
d.cell(nr,1,"HOW TO COVER 10 CATEGORIES ON 4 TOURNEYS + 3 DRAFTS A DAY").font=Font(name=F,bold=True,size=12)
lines=[
 "FEASIBILITY: you need 1,523 pts over 35 days = 43.5/day across 10 categories.",
 "   In PTCS 5 you produced 43.9/day across just 6. THE CAPACITY IS ALREADY THERE.",
 "   This is a REDISTRIBUTION problem, not an effort problem.",
 "",
 "PTCS 5 OVERKILL — 418 points spent above the cutoff:",
 "      Diamond  283 vs  81 needed   = +202 wasted",
 "      PD Daily 403 vs 267 needed   = +136 wasted",
 "      Gold/Cap/Live                = +80 combined",
 "   That Diamond surplus alone is roughly Iron + Bronze targets COMBINED.",
 "",
 "THE RULE: once a category clears its target, STOP feeding it and move those entries to a laggard.",
 "   Check the Gap column daily. Anything at 0 = redirect.",
 "",
 "SLOTS: the 4-tournament cap is CONCURRENT, not daily. Events fire on fixed times all day and free",
 "   the slot, so you can run 10-15 tournaments/day. Keep all 4 slots full at all times — an idle",
 "   slot is the only real waste.",
 "",
 "MULTI-TAG EVENTS still do the heavy lifting:",
 "   • 'Daily Live <Tier>'   -> that Tier + LIVE   (Live Iron/Bronze/Silver/Gold/Diamond/Open)",
 "   • '<Tier> Cap' / 'Slots'-> that Tier + CAP    (Iron Cap, Bronze Cap, Silver Slots, Gold Slots, Diamond Slots, Open High Cap)",
 "   • Live DRAFTS           -> PD Daily + LIVE    (Live Breakfast / Speed Run / All Night / Don't Forget)",
 "   -> Live and Cap should never need a dedicated slot; they ride along on everything.",
 "",
 "DAILY SHAPE: keep 4 slots full, rotating tiers so Iron/Bronze/Silver/Gold/Diamond/Open each get",
 "   2-3 entries a day. Prefer the Live-<tier> and <tier>-Cap variants over the plain version -",
 "   same effort, two categories.",
 "",
 "LESSONS FROM PTCS 5 (5 of 6 berths):",
 "   • Start every category DAY ONE - Silver missed by 16 only because it began in week 3.",
 "   • Points are convex: a WIN in a 64 = 25 pts, 17th-32nd = 1. Deep runs >> volume.",
 "   • Best EV events: Diamonds are Forever 4.46/entry, Diamond & Friends Slots 3.93, Daily Diamond 2.77.",
 "   • You convert better in 64-team fields (2.72/entry vs 1.75 in 128s), but 128s pay ~50% more at",
 "     the top - take 128s when you NEED a big score, 64s when accumulating.",
 "   • Fields do NOT soften late in a period (verified across 6,175 tournaments) - no reason to wait.",
]
for i,t in enumerate(lines):
    d.cell(nr+1+i,1,("• "+t) if t and not t.startswith("   ") and not t.endswith(":") else t).font=Font(name=F,size=10,color="444444")
d.freeze_panes="A5"

# ============ PTCS -> PTMS -> PTWC championship ladder ============
p=wb.create_sheet("PTCS-PTMS")
p["A1"]="The Championship Ladder — PTCS → PTMS → PTWC"; p["A1"].font=title
p["A2"]="Separate from the 10 category standings. Source: 'PT 27 Road to the Perfect Team World Championship'."; p["A2"].font=sub

p["A4"]="POINTS EARNED FROM A PTCS FINISH (feeds PTMS + PTWC qualifying)"; p["A4"].font=Font(name=F,bold=True,size=12)
pt=[("1st",300),("2nd",100),("3rd-4th",75),("5th-8th",40),("9th-16th",20),("17th-32nd",5),("33rd-64th",2),("65th-128th",1)]
for c,h in enumerate(["Finish","Points"],1):
    x=p.cell(5,c,h); x.font=hfont; x.fill=hf; x.alignment=ctr; x.border=bd
for i,(f_,v) in enumerate(pt):
    p.cell(6+i,1,f_).font=bf; p.cell(6+i,2,v).font=bf
    for c in (1,2): p.cell(6+i,c).border=bd; p.cell(6+i,c).alignment=ctr

p["D4"]="YOUR PTCS 5 (Aug 8) — all 128-team fields"; p["D4"].font=Font(name=F,bold=True,size=12)
hdr=["Event","Finish","Points","Standing"]
for c,h in enumerate(hdr,4):
    x=p.cell(5,c,h); x.font=hfont; x.fill=hf; x.alignment=ctr; x.border=bd
P5=[("PTCS Gold - Period 5","17th-32nd",5,"Tournament"),
    ("PTCS Daily PD - Period 5","17th-32nd",5,"Perfect Draft"),
    ("PTCS Diamond - Period 5","33rd-64th",2,"Tournament"),
    ("PTCS Cap - Period 5","65th-128th",1,"Tournament"),
    ("PTCS Live - Period 5","65th-128th",1,"Tournament")]
for i,row in enumerate(P5):
    for c,v in enumerate(row,4):
        x=p.cell(6+i,c,v); x.font=bf; x.border=bd; x.alignment=left if c==4 else ctr
p.cell(11,4,"PTCS 5 TOTAL").font=Font(name=F,bold=True,size=11)
p.cell(11,6,"=SUM(F6:F10)").font=Font(name=F,bold=True,size=11)
for c in range(4,8): p.cell(11,c).border=bd

p["A16"]="CUMULATIVE PTCS STANDINGS — top 128 in each qualifies for the PTMS"; p["A16"].font=Font(name=F,bold=True,size=12)
hdr2=["PTCS","Date","Tournament pts","Perfect Draft pts","Total","Notes"]
for c,h in enumerate(hdr2,1):
    x=p.cell(17,c,h); x.font=hfont; x.fill=hf; x.alignment=ctr; x.border=bd
HIST=[("PTCS 1","Apr 11",0,0,0,"did not qualify"),
      ("PTCS 2","May 9",0,0,0,"did not qualify"),
      ("PTCS 3","Jun 6",1,0,1,"1 berth (Diamond) - 65th-128th"),
      ("PTCS 4","Jul 11",20,0,20,"1 berth (Diamond) - 9th-16th. BEST PTCS RESULT SO FAR"),
      ("PTCS 5","Aug 8",9,5,14,"5 berths - Gold 17-32 (5), Diamond 33-64 (2), Cap + Live 65-128 (1 ea) = 9 tourney; Daily PD 17-32 = 5"),
      ("TOTAL","","=SUM(C18:C22)","=SUM(D18:D22)","=SUM(E18:E22)","top 128 of EACH standing makes the PTMS")]
for i,row in enumerate(HIST):
    for c,v in enumerate(row,1):
        x=p.cell(18+i,c,v); x.font=Font(name=F,bold=True,size=11) if row[0].startswith("KNOWN") else bf
        x.border=bd; x.alignment=left if c in (1,6) else ctr

p["A25"]="THE TABLE IS VIOLENTLY CONVEX — BERTHS GET YOU IN, DEPTH GETS YOU POINTS"; p["A25"].font=Font(name=F,bold=True,size=12)
conv=[
 "PTCS 4: ONE berth, finished 9th-16th  ->  20 pts",
 "PTCS 5: FIVE berths, best was 17th-32nd  ->  14 pts",
 "   Five times the entries produced FEWER points. The jump from 17th-32nd (5) to 9th-16th (20) is 4x;",
 "   9th-16th to 5th-8th is another 2x. A single quarterfinal (40) beats eight 17th-32nd finishes.",
 "   Berths are the price of admission. Points come from depth, not breadth.",
]
for i,t in enumerate(conv): p.cell(26+i,1,t).font=Font(name=F,size=10,color="444444")

p["A33"]="WHICH WORLD CHAMPIONSHIP YOUR WORK FEEDS"; p["A33"].font=Font(name=F,bold=True,size=12)
notes=[
 "PTWC 1 (Aug 29-30) qualifies from:  League play to Aug 23  +  PT Live pts/50  +  Weekly leaderboard pts to Aug 2  +  PTCS 1-5  +  PTMS 1-2",
 "PTWC 2 (Feb 21-22) qualifies from:  League play Aug 24 on  +  PT Live pts/50  +  Weekly leaderboard pts Aug 4-Jan 31  +  PTCS 6-11  +  PTMS 3-4",
 "",
 ">> PTCS 5 was the LAST input to PTWC 1. Everything you score in PTCS 6 feeds PTWC 2 instead.",
 ">> PTMS 2 is Aug 22 and its field is set by the PTCS 1-5 standings, which are now final.",
 "   Nothing you enter between now and then changes whether you're in it.",
 "",
 "Cap and Live standings points are excluded from PTWC qualifying (no double-dipping) - they always",
 "ride along with a tier. The PTCS Cap and PTCS Live EVENTS are separate berths you actually played,",
 "so they should still count toward the PTMS tournament standing. Worth confirming before relying on it.",
 "",
 "Doc inconsistency: page 1 dates PTWC 1 as Aug 29-30, page 3 says Aug 30-31.",
]
for i,t in enumerate(notes):
    p.cell(34+i,1,t).font=Font(name=F,size=10,color="444444")
for c,w in enumerate([15,10,16,18,10,60],1): p.column_dimensions[get_column_letter(c)].width=w
p.freeze_panes="A5"

import sys
wb.save(sys.argv[1] if len(sys.argv)>1 else "PTCS6_tracker.xlsx")
print("built all-10")
