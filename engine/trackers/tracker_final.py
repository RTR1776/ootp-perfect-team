from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
F="Arial"
hf=PatternFill("solid",fgColor="1F3B57"); hfont=Font(name=F,bold=True,color="FFFFFF",size=11)
title=Font(name=F,bold=True,size=15); sub=Font(name=F,italic=True,color="666666",size=10)
bf=Font(name=F,size=11); gray=Font(name=F,size=10,color="888888")
thin=Side(style="thin",color="CCCCCC"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal="center"); left=Alignment(horizontal="left")
GRN=PatternFill("solid",fgColor="D6F5E3"); AMB=PatternFill("solid",fgColor="FFF3CD")
RED=PatternFill("solid",fgColor="FFE1E1"); dim=PatternFill("solid",fgColor="F2F4F7")

# OFFICIAL period standings thru Jul 26 | rank | 128th now | projected final (x28/21)
CATS=[  # name, official, rank, line_now, proj_final, chased
 ("Diamond",244,11,61,81,True),("PDD (D.Daily)",356,28,200,267,True),
 ("Gold",119,64,68,91,True),("Live",134,87,106,141,True),
 ("Cap",127,103,107,143,True),("Silver",30,286,77,103,True),
 ("Open",7,493,77,103,True),("W. Draft",6,735,58,77,False),
]
COL={"Diamond":"C","Live":"D","Cap":"E","Gold":"F","Silver":"G","PDD (D.Daily)":"H","Open":"I","W. Draft":"J"}
# final-week log: [Date,Day,Dia,Live,Cap,Gold,Silver,PDD,Open,WDraft,Notes]
DAYS=[
 ["Mon Jul 27","Mon",10,0,17,0,6,0,7,0,"Dia Slots 3/4 +10C; Gold Floor Cap 9-16 +6 Open/Cap; Late Silver 9-16 +6"],
 ["Tue Jul 28","Tue",3,1,3,0,7,0,0,1,"Silver Heart 9-16 +3; Late Silver 17-32 +3; Slamboree +1; Dia Slots 9-16 +3C"],
 ["Wed Jul 29","Wed",12,5,0,0,3,11,0,0,"Wed Ice 17-32 +6D; Dia Heart 9-16 +6D; Speed Run 9-16 +3L/+3P; Laptophound 9-16 +6P"],
 ["Thu Jul 30","Thu",7,0,12,0,7,3,0,0,"Dia Slots 5-8 +6D/+6C; Silver&Friends 9-16 +6S/+6C; Slamboree +1S; Busy 9-16 +3P"],
 ["Fri Jul 31","Fri",0,0,3,0,6,1,0,10,"Late Silver 17-32 +3S; Friday Nightmare Cap 17-32 (128) +3S/+3C; PD Weekly 9-16 +10 WD"],
 ["Fri Jul 31 (pm)","Fri",0,1,1,0,2,0,0,0,"Live Silver 33-64 (128) +1L/+1S; Silver Slots 17-32 +1S/+1C"],
 ["Sat Aug 1","Sat",6,21,4,0,19,25,0,10,"Late Silver semi 3/4 +15S; Live All Night 2nd +15L/+15P; Breakfast 5-8 +6L/+6P; Slamboree 17-32 +3S"],
 ["Sun Aug 2","Sun",1,3,4,0,7,7,0,0,"FINAL DAY - Silver Heart 9-16 +3S; Deadball Rd3 +3S/+3C; Slamboree Rd2 +1S; Laptophound +3P"],
]
wb=Workbook()
log=wb.active; log.title="Final Week Log"
log["A1"]="PTCS 5 — FINAL WEEK (Jul 27 – Aug 2)"; log["A1"].font=title
log["A2"]="Base = OFFICIAL period standings thru Jul 26. Log only new points from Jul 27."; log["A2"].font=sub
cols=["Date","Day","Diamond","Live","Cap","Gold","Silver","PDD","Open","W.Draft","Notes"]
for c,h in enumerate(cols,1):
    x=log.cell(4,c,h); x.font=hfont; x.fill=hf; x.alignment=ctr; x.border=bd
r=5
for row in DAYS:
    for c,v in enumerate(row,1):
        x=log.cell(r,c,v); x.font=bf; x.border=bd; x.alignment=left if c in(1,2,11) else ctr
    r+=1
for _ in range(12):
    for c in range(1,12): log.cell(r,c,"").border=bd
    r+=1
LAST=r-1
for c,w in enumerate([12,6,9,8,8,8,8,8,8,9,62],1): log.column_dimensions[get_column_letter(c)].width=w
log.freeze_panes="C5"

d=wb.create_sheet("Dashboard",0)
d["A1"]="PTCS 5 — Final Week Push"; d["A1"].font=title
d["A2"]="Base = OFFICIAL standings thru Jul 26 (all 8 categories). Projected final cutoff = current 128th x 28/21 days. Period ends Aug 2."; d["A2"].font=sub
head=["Category","Official","Rank","128th now","Proj final cutoff","Added","Total","Gap","Per day (2d)","Verdict"]
for c,h in enumerate(head,1):
    x=d.cell(4,c,h); x.font=hfont; x.fill=hf; x.alignment=ctr; x.border=bd
rr=5
for name,off,rank,now,proj,chase in CATS:
    d.cell(rr,1,name).font=Font(name=F,bold=True,size=11) if chase else Font(name=F,size=11,color="777777")
    d.cell(rr,2,off).font=bf
    d.cell(rr,3,rank).font=gray
    d.cell(rr,4,now).font=gray
    d.cell(rr,5,proj).font=bf
    d.cell(rr,6,f"=SUM('Final Week Log'!{COL[name]}5:{COL[name]}{LAST})").font=bf
    d.cell(rr,7,f"=B{rr}+F{rr}").font=Font(name=F,bold=True,size=11)
    d.cell(rr,8,f"=MAX(0,E{rr}-G{rr})").font=bf
    d.cell(rr,9,f"=ROUND(H{rr}/2,1)").font=bf
    d.cell(rr,10,f'=IF(G{rr}>=E{rr},"CLEAR","+"&TEXT(H{rr},"0")&" needed")')
    for c in range(1,11):
        x=d.cell(rr,c); x.border=bd
        if c not in (1,10): x.alignment=ctr
        if not chase: x.fill=dim
    rr+=1
for c,w in enumerate([15,9,7,10,17,8,8,8,12,16],1): d.column_dimensions[get_column_letter(c)].width=w
nr=rr+1
d.cell(nr,1,"Final week priorities").font=Font(name=F,bold=True,size=12)
for i,t in enumerate([
 "CLEAR already: Diamond (244 vs 81 proj — rank 11), PDD (356 vs 267 — rank 28), Gold (119 vs 91).",
 "LIVE is the near-lock: +7 to the projected line. One Live-draft finish does it. Don't leave this one.",
 "CAP is the live fight: +16 after Monday's big +17. Your Diamond+Cap and Open+Cap games are carrying it.",
 "SILVER +73 and OPEN +96 in 6 days = ~12-16/day each. Reachable only if you go hard on Silver+Cap / Open+Cap games (they triple-dip).",
 "Best EV per your 368-entry history: Diamonds are Forever 4.46, Diamond & Friends Slots 3.93 (also Cap), Daily Diamond 2.77.",
 "You score better in 64-team fields (2.72/entry) than 128s (1.75) — prefer 64s when both are open.",
 "Not chasing Bronze/Iron (per LJ).",
]): d.cell(nr+1+i,1,"• "+t).font=Font(name=F,size=10,color="444444")
d.freeze_panes="A5"
wb.save("/sessions/stoic-sleepy-goodall/mnt/outputs/PTCS5_tracker.xlsx")
print("built")
